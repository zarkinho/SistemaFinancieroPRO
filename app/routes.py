# ============================================================
# PROYECTO : Sistema Financiero PRO
# ARCHIVO  : routes.py
# VERSIÓN  : 5.0
# ============================================================

# ============================================================
# IMPORTACIONES
# ============================================================

from flask import (
    Blueprint,
    render_template,
    redirect,
    url_for,
    flash,
    jsonify,
    send_file,
    request
)

from sqlalchemy import func

from openpyxl import Workbook
from openpyxl.styles import Font

import os

from app import db

from app.forms import MovimientoForm

from app.models import (
    Cuenta,
    Categoria,
    Movimiento,
    Presupuesto,
    Configuracion
)

from app.reportes_pdf import generar_pdf

# ============================================================
# BLUEPRINT
# ============================================================

main = Blueprint(
    "main",
    __name__
)

# ============================================================
# FUNCIONES AUXILIARES
# ============================================================

def calcular_totales():

    ingresos = db.session.query(

        func.coalesce(
            func.sum(Movimiento.valor),
            0
        )

    ).filter(

        Movimiento.tipo == "INGRESO"

    ).scalar()

    gastos = db.session.query(

        func.coalesce(
            func.sum(Movimiento.valor),
            0
        )

    ).filter(

        Movimiento.tipo == "GASTO"

    ).scalar()

    inversiones = db.session.query(

        func.coalesce(
            func.sum(Movimiento.valor),
            0
        )

    ).filter(

        Movimiento.tipo == "INVERSION"

    ).scalar()

    saldo = ingresos - gastos - inversiones

    return {

        "saldo": saldo,

        "ingresos": ingresos,

        "gastos": gastos,

        "inversiones": inversiones

    }


def calcular_saldo_cuenta(cuenta):

    ingresos = db.session.query(

        func.coalesce(
            func.sum(Movimiento.valor),
            0
        )

    ).filter(

        Movimiento.cuenta_id == cuenta.id,

        Movimiento.tipo == "INGRESO"

    ).scalar()

    gastos = db.session.query(

        func.coalesce(
            func.sum(Movimiento.valor),
            0
        )

    ).filter(

        Movimiento.cuenta_id == cuenta.id,

        Movimiento.tipo == "GASTO"

    ).scalar()

    inversiones = db.session.query(

        func.coalesce(
            func.sum(Movimiento.valor),
            0
        )

    ).filter(

        Movimiento.cuenta_id == cuenta.id,

        Movimiento.tipo == "INVERSION"

    ).scalar()

    return (

        cuenta.saldo_inicial +
        ingresos -
        gastos -
        inversiones

    )


# ============================================================
# DASHBOARD
# ============================================================

@main.route("/")
def dashboard():

    totales = calcular_totales()

    movimientos = (

        Movimiento.query

        .order_by(

            Movimiento.fecha.desc()

        )

        .limit(10)

        .all()

    )

    cuentas = Cuenta.query.order_by(

        Cuenta.nombre

    ).all()

    saldos_cuentas = []

    for cuenta in cuentas:

        saldos_cuentas.append(

            {

                "nombre": cuenta.nombre,

                "saldo": calcular_saldo_cuenta(cuenta)

            }

        )

    mayor_cuenta = None
    menor_cuenta = None

    if saldos_cuentas:

        mayor_cuenta = max(

            saldos_cuentas,

            key=lambda c: c["saldo"]

        )

        menor_cuenta = min(

            saldos_cuentas,

            key=lambda c: c["saldo"]

        )

    if totales["ingresos"] > 0:

        porcentaje_gastos = round(

            (

                totales["gastos"]

                /

                totales["ingresos"]

            ) * 100,

            1

        )

    else:

        porcentaje_gastos = 0

    return render_template(

        "dashboard.html",

        saldo=totales["saldo"],

        ingresos=totales["ingresos"],

        gastos=totales["gastos"],

        inversiones=totales["inversiones"],

        movimientos=movimientos,

        saldos_cuentas=saldos_cuentas,

        mayor_cuenta=mayor_cuenta,

        menor_cuenta=menor_cuenta,

        porcentaje_gastos=porcentaje_gastos,

        datos_grafico=[

            totales["ingresos"],

            totales["gastos"],

            totales["inversiones"]

        ]

    )

# ============================================================
# NUEVO MOVIMIENTO
# ============================================================

@main.route("/nuevo", methods=["GET", "POST"])
def nuevo_movimiento():

    form = MovimientoForm()

    cuentas = Cuenta.query.order_by(
        Cuenta.nombre
    ).all()

    form.cuenta.choices = [
        (cuenta.id, cuenta.nombre)
        for cuenta in cuentas
    ]

    tipo = form.tipo.data or "INGRESO"

    categorias = Categoria.query.filter_by(
        tipo=tipo
    ).order_by(
        Categoria.nombre
    ).all()

    form.categoria.choices = [
        (categoria.id, categoria.nombre)
        for categoria in categorias
    ]

    if form.validate_on_submit():

        movimiento = Movimiento(

            tipo=form.tipo.data,

            descripcion=form.descripcion.data,

            valor=form.valor.data,

            observaciones=form.observaciones.data,

            cuenta_id=form.cuenta.data,

            categoria_id=form.categoria.data

        )

        db.session.add(movimiento)

        db.session.commit()

        flash(
            "Movimiento registrado correctamente.",
            "success"
        )

        return redirect(
            url_for("main.dashboard")
        )

    return render_template(
        "nuevo_movimiento.html",
        form=form
    )


# ============================================================
# EDITAR MOVIMIENTO
# ============================================================

@main.route("/editar/<int:id>", methods=["GET", "POST"])
def editar_movimiento(id):

    movimiento = Movimiento.query.get_or_404(id)

    form = MovimientoForm(obj=movimiento)

    cuentas = Cuenta.query.order_by(
        Cuenta.nombre
    ).all()

    form.cuenta.choices = [
        (cuenta.id, cuenta.nombre)
        for cuenta in cuentas
    ]

    categorias = Categoria.query.filter_by(
        tipo=movimiento.tipo
    ).order_by(
        Categoria.nombre
    ).all()

    form.categoria.choices = [
        (categoria.id, categoria.nombre)
        for categoria in categorias
    ]

    if form.validate_on_submit():

        movimiento.tipo = form.tipo.data
        movimiento.descripcion = form.descripcion.data
        movimiento.valor = form.valor.data
        movimiento.observaciones = form.observaciones.data
        movimiento.cuenta_id = form.cuenta.data
        movimiento.categoria_id = form.categoria.data

        db.session.commit()

        flash(
            "Movimiento actualizado correctamente.",
            "success"
        )

        return redirect(
            url_for("main.dashboard")
        )

    form.cuenta.data = movimiento.cuenta_id
    form.categoria.data = movimiento.categoria_id

    return render_template(
        "nuevo_movimiento.html",
        form=form
    )


# ============================================================
# ELIMINAR MOVIMIENTO
# ============================================================

@main.route("/eliminar/<int:id>")
def eliminar_movimiento(id):

    movimiento = Movimiento.query.get_or_404(id)

    db.session.delete(movimiento)

    db.session.commit()

    flash(
        "Movimiento eliminado correctamente.",
        "success"
    )

    return redirect(
        url_for("main.dashboard")
    )


# ============================================================
# HISTORIAL
# ============================================================

@main.route("/historial")
def historial():

    movimientos = (

        Movimiento.query

        .order_by(
            Movimiento.fecha.desc()
        )

        .all()

    )

    return render_template(

        "historial.html",

        movimientos=movimientos

    )


# ============================================================
# CARGAR CATEGORÍAS (AJAX)
# ============================================================

@main.route("/categorias/<tipo>")
def obtener_categorias(tipo):

    categorias = Categoria.query.filter_by(

        tipo=tipo

    ).order_by(

        Categoria.nombre

    ).all()

    return jsonify(

        [

            {

                "id": categoria.id,

                "nombre": categoria.nombre

            }

            for categoria in categorias

        ]

    )

# ============================================================
# ESTADÍSTICAS
# ============================================================

@main.route("/estadisticas")
def estadisticas():

    totales = calcular_totales()

    return render_template(

        "estadisticas.html",

        saldo=totales["saldo"],

        ingresos=totales["ingresos"],

        gastos=totales["gastos"],

        inversiones=totales["inversiones"]

    )


# ============================================================
# REPORTES FINANCIEROS
# ============================================================

@main.route("/reportes")
def reportes():

    totales = calcular_totales()

    movimientos = (

        Movimiento.query

        .order_by(

            Movimiento.fecha.desc()

        )

        .all()

    )

    return render_template(

        "reportes.html",

        saldo=totales["saldo"],

        ingresos=totales["ingresos"],

        gastos=totales["gastos"],

        inversiones=totales["inversiones"],

        movimientos=movimientos

    )


# ============================================================
# EXPORTAR MOVIMIENTOS A EXCEL
# ============================================================

@main.route("/exportar_excel")
def exportar_excel():

    libro = Workbook()

    hoja = libro.active

    hoja.title = "Movimientos"

    encabezados = [

        "Fecha",
        "Tipo",
        "Cuenta",
        "Categoría",
        "Descripción",
        "Valor"

    ]

    for columna, texto in enumerate(

        encabezados,

        start=1

    ):

        celda = hoja.cell(

            row=1,

            column=columna

        )

        celda.value = texto

        celda.font = Font(

            bold=True

        )

    movimientos = (

        Movimiento.query

        .order_by(

            Movimiento.fecha.desc()

        )

        .all()

    )

    fila = 2

    for movimiento in movimientos:

        hoja.cell(

            fila,

            1

        ).value = movimiento.fecha.strftime(

            "%d/%m/%Y"

        )

        hoja.cell(

            fila,

            2

        ).value = movimiento.tipo

        hoja.cell(

            fila,

            3

        ).value = movimiento.cuenta.nombre

        hoja.cell(

            fila,

            4

        ).value = movimiento.categoria.nombre

        hoja.cell(

            fila,

            5

        ).value = movimiento.descripcion

        hoja.cell(

            fila,

            6

        ).value = movimiento.valor

        fila += 1

    ruta = os.path.join(

        os.getcwd(),

        "movimientos.xlsx"

    )

    libro.save(

        ruta

    )

    return send_file(

        ruta,

        as_attachment=True

    )


# ============================================================
# EXPORTAR REPORTE PDF
# ============================================================

@main.route("/exportar_pdf")
def exportar_pdf():

    return generar_pdf()

# ============================================================
# CONFIGURACIÓN DEL SISTEMA
# ============================================================

@main.route("/configuracion", methods=["GET", "POST"])
def configuracion():

    configuracion = Configuracion.query.first()

    if configuracion is None:

        configuracion = Configuracion()

        db.session.add(configuracion)

        db.session.commit()

    if request.method == "POST":

        configuracion.nombre_empresa = request.form.get(
            "nombre_empresa"
        )

        configuracion.propietario = request.form.get(
            "propietario"
        )

        configuracion.telefono = request.form.get(
            "telefono"
        )

        configuracion.correo = request.form.get(
            "correo"
        )

        configuracion.direccion = request.form.get(
            "direccion"
        )

        configuracion.ciudad = request.form.get(
            "ciudad"
        )

        configuracion.moneda = request.form.get(
            "moneda"
        )

        configuracion.color_principal = request.form.get(
            "color_principal"
        )

        configuracion.color_secundario = request.form.get(
            "color_secundario"
        )

        db.session.commit()

        flash(
            "Configuración guardada correctamente.",
            "success"
        )

        return redirect(
            url_for("main.configuracion")
        )

    return render_template(

        "configuracion.html",

        configuracion=configuracion

    )


# ============================================================
# ACERCA DE
# ============================================================

@main.route("/acerca")
def acerca():

    return redirect(

        url_for("main.dashboard")

    )


# ============================================================
# LIMPIAR CATEGORÍAS DUPLICADAS
# ============================================================

@main.route("/utilidades/reparar-categorias")
def reparar_categorias():

    categorias = Categoria.query.order_by(

        Categoria.id

    ).all()

    vistos = set()

    eliminadas = 0

    for categoria in categorias:

        clave = (

            categoria.nombre.strip().upper(),

            categoria.tipo

        )

        if clave in vistos:

            db.session.delete(

                categoria

            )

            eliminadas += 1

        else:

            vistos.add(

                clave

            )

    db.session.commit()

    flash(

        f"Se eliminaron {eliminadas} categorías duplicadas.",

        "success"

    )

    return redirect(

        url_for("main.dashboard")

    )


# ============================================================
# ERROR 404
# ============================================================

@main.app_errorhandler(404)
def pagina_no_encontrada(error):

    flash(

        "La página solicitada no existe.",

        "warning"

    )

    return redirect(

        url_for("main.dashboard")

    )


# ============================================================
# FIN DEL ARCHIVO
# ============================================================